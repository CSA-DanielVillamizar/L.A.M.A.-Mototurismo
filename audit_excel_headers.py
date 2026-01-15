import openpyxl
import sys

# Leer el Excel real
excel_path = r"c:\Users\DanielVillamizar\COR L.A.MA\INSUMOS\(COL) PEREIRA CORTE NACIONAL.xlsx"
wb = openpyxl.load_workbook(excel_path)

# Buscar la sheet ODOMETER
for sheet_name in wb.sheetnames:
    print(f"\n=== SHEET: {sheet_name} ===")
    ws = wb[sheet_name]
    
    # Leer fila 7 (header)
    if sheet_name.upper() == "ODOMETER":
        print("\nFila 7 (HEADERS):")
        for col_idx, cell in enumerate(ws[7], 1):
            if cell.value:
                header = str(cell.value)
                # Mostrar el header y su primer carácter (para detectar espacios)
                first_char_code = ord(header[0]) if header else None
                print(f"  Col {col_idx:2d}: '{header}' (primer char ASCII: {first_char_code})")
                
                # Si tiene espacio inicial
                if header.startswith(' '):
                    print(f"           ⚠️  ESPACIO INICIAL DETECTADO")

wb.close()
