import openpyxl
from pathlib import Path

file_path = Path('INSUMOS/(COL) PEREIRA CORTE NACIONAL.xlsx')

print("=" * 70)
print("EXTRAYENDO VALIDACIÓN DE DATOS - COLUMNA STATUS")
print("=" * 70)

wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['ODOMETER']

# La columna STATUS debería ser la columna G (índice 7)
# Buscar la validación de datos en esa columna
status_col_letter = None

# Primero encontrar qué columna es STATUS
header_row = 8  # Fila 8 contiene los headers
for cell in ws[header_row]:
    if cell.value and 'STATUS' in str(cell.value).upper():
        status_col_letter = cell.column_letter
        print(f'\nColumna STATUS encontrada: {status_col_letter}')
        break

if status_col_letter:
    # Obtener validaciones de datos
    print(f'\n=== VALIDACIONES DE DATOS EN COLUMNA {status_col_letter} ===')
    
    # Buscar en data validations
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            # Verificar si esta validación aplica a la columna STATUS
            ranges = str(dv.sqref)
            if status_col_letter in ranges:
                print(f'\nRango: {ranges}')
                print(f'Tipo: {dv.type}')
                
                if dv.type == 'list':
                    if dv.formula1:
                        print(f'Lista de valores:')
                        # Limpiar la fórmula (viene como "valor1,valor2,...")
                        values = dv.formula1.strip('"').split(',')
                        for i, val in enumerate(values, 1):
                            print(f'  {i:2d}. {val.strip()}')
                    
                    if dv.showDropDown is not None:
                        print(f'Mostrar dropdown: {not dv.showDropDown}')
                
                print(f'Permitir blancos: {dv.allow_blank}')
                if dv.errorTitle:
                    print(f'Error título: {dv.errorTitle}')
                if dv.error:
                    print(f'Error mensaje: {dv.error}')
    else:
        print('\nNo se encontraron validaciones de datos en el archivo.')
    
    # Verificar si hay valores únicos en la columna
    print(f'\n=== VALORES ÚNICOS EN LA COLUMNA (primeras 100 filas) ===')
    status_values = set()
    for row in range(9, 109):  # Filas de datos
        cell_value = ws[f'{status_col_letter}{row}'].value
        if cell_value:
            status_values.add(str(cell_value).strip())
    
    for i, val in enumerate(sorted(status_values), 1):
        print(f'  {i:2d}. {val}')

else:
    print('\n[ERROR] No se encontró la columna STATUS')

wb.close()
print("\n" + "=" * 70)
